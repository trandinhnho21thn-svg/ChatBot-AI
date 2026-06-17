"""
Tạo file Excel mẫu cho chatbot Rapido Bảo Hành
Cấu trúc 3 sheet:
  1. HƯỚNG DẪN  - giải thích cách dùng
  2. SẢN PHẨM   - thông tin chính mỗi model
  3. MÃ LỖI     - danh sách lỗi theo model
  4. LINH KIỆN  - danh sách linh kiện theo model
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Màu sắc ──────────────────────────────────────────────
C_HEADER_BLUE   = "1E40AF"   # header sản phẩm
C_HEADER_RED    = "BE123C"   # header mã lỗi
C_HEADER_GREEN  = "065F46"   # header linh kiện
C_HEADER_GRAY   = "374151"   # header hướng dẫn
C_ROW_ALT       = "F0F9FF"   # dòng xen kẽ
C_ROW_ALT2      = "FFF1F2"   # dòng xen kẽ lỗi
C_ROW_ALT3      = "ECFDF5"   # dòng xen kẽ linh kiện
C_YELLOW        = "FEF9C3"   # highlight cột bắt buộc
C_GUIDE_BG      = "1E3A5F"   # nền tối hướng dẫn
C_WHITE         = "FFFFFF"
C_LIGHT_BLUE    = "DBEAFE"
C_LIGHT_RED     = "FFE4E6"
C_LIGHT_GREEN   = "D1FAE5"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Segoe UI")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_cell(cell, bg_hex, text, bold=True, size=11):
    cell.value = text
    cell.fill = fill(bg_hex)
    cell.font = font(bold=bold, color=C_WHITE, size=size)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = thin_border()

def style_data_cell(cell, value="", bg_hex=C_WHITE, bold=False,
                    wrap=False, align_h="left", color="1E293B"):
    cell.value = value
    cell.fill = fill(bg_hex)
    cell.font = font(bold=bold, color=color, size=10)
    cell.alignment = align(align_h, "center", wrap=wrap)
    cell.border = thin_border()

# ═══════════════════════════════════════════════════════════
wb = Workbook()

# ─── SHEET 1: HƯỚNG DẪN ────────────────────────────────
ws_guide = wb.active
ws_guide.title = "📋 HƯỚNG DẪN"
ws_guide.sheet_view.showGridLines = False
ws_guide.column_dimensions["A"].width = 4
ws_guide.column_dimensions["B"].width = 28
ws_guide.column_dimensions["C"].width = 70

# Tiêu đề
ws_guide.merge_cells("B1:C1")
c = ws_guide["B1"]
c.value = "🔧 HƯỚNG DẪN SỬ DỤNG FILE MẪU CHATBOT RAPIDO BẢO HÀNH"
c.fill = fill(C_GUIDE_BG)
c.font = Font(bold=True, color=C_WHITE, size=15, name="Segoe UI")
c.alignment = align("center", "center")
ws_guide.row_dimensions[1].height = 40

ws_guide.merge_cells("B2:C2")
c = ws_guide["B2"]
c.value = "Điền thông tin vào các sheet tương ứng, sau đó tải lên chatbot qua nút ↑ Import"
c.fill = fill("334155")
c.font = font(color=C_WHITE, size=10, italic=True)
c.alignment = align("center", "center")
ws_guide.row_dimensions[2].height = 22

rows_guide = [
    ("",  "",  ""),
    ("",  "📌 CÁC SHEET TRONG FILE",  ""),
    ("",  "Sheet 2 – SẢN PHẨM",       "Nhập thông tin cơ bản mỗi model: loại SP, tên model, đặc điểm, hướng dẫn SD, link ảnh."),
    ("",  "Sheet 3 – MÃ LỖI",          "Nhập các mã lỗi của từng model. Mỗi mã lỗi = 1 dòng. Dùng đúng Tên Model ở cột A."),
    ("",  "Sheet 4 – LINH KIỆN",       "Nhập linh kiện thay thế. Mỗi linh kiện = 1 dòng. Dùng đúng Tên Model ở cột A."),
    ("",  "",  ""),
    ("",  "📌 LƯU Ý QUAN TRỌNG",  ""),
    ("",  "Cột bắt buộc",              "Tên Model (cột Tên Model / Model) là cột BẮT BUỘC. Nếu để trống sẽ bị bỏ qua."),
    ("",  "Phân biệt hoa/thường",      "Tên Model phải KHỚP CHÍNH XÁC giữa sheet SẢN PHẨM và sheet MÃ LỖI / LINH KIỆN."),
    ("",  "Cột Link ảnh",             "Dán URL hình ảnh (http://... hoặc https://...). Để trống nếu dùng gắn ảnh từ máy tính."),
    ("",  "Nhiều mã lỗi / linh kiện", "Thêm nhiều dòng cho cùng một model ở sheet MÃ LỖI hoặc LINH KIỆN."),
    ("",  "Khi import vào chatbot",    "Nhấn nút ↑ Upload → chọn file Excel này → chọn 'Thêm vào' hoặc 'Ghi đè'."),
    ("",  "",  ""),
    ("",  "📌 CÁCH GẮN ẢNH TỪ MÁY TÍNH",  ""),
    ("",  "Bước 1",                    "Import file Excel này vào chatbot (nút ↑ Upload)."),
    ("",  "Bước 2",                    "Nhấn nút 🖼 (màu tím) ở góc phải mục Danh Sách Model → hộp thoại 'Gắn Ảnh Hàng Loạt' mở ra."),
    ("",  "Bước 3 – Tự động ghép",     "Nhấn 'Tải nhiều ảnh cùng lúc' → chọn tất cả ảnh trong máy. Hệ thống TỰ GHÉP theo tên file.\nVD: file 'RI-2000ES.jpg' → tự gắn vào model 'RI-2000ES'."),
    ("",  "Bước 3 – Thủ công",         "Hoặc nhấn 'Chọn ảnh SP' / 'Chọn ảnh HD' từng dòng để gắn thủ công."),
    ("",  "Ảnh hướng dẫn (HD)",        "Đặt tên file chứa 'hd', 'guide' hoặc 'huongdan' để tự ghép vào ảnh hướng dẫn.\nVD: 'RI-2000ES-hd.jpg' → ghép vào ảnh HD của RI-2000ES."),
    ("",  "Nén ảnh tự động",           "Ảnh được nén thông minh (WebP/JPEG chất lượng cao 90%). Chỉ resize nếu ảnh vượt 1600px.\nKhông vỡ, không nhòe khi phóng to."),
    ("",  "",  ""),
    ("",  "📌 VÍ DỤ MÃ LỖI",  ""),
    ("",  "Cột Mã lỗi",                "E1, E2, F1, H01, ... (gõ ngắn gọn để chatbot dễ nhận diện)"),
    ("",  "Cột Nguyên nhân",           "Mô tả nguyên nhân gây ra lỗi. VD: Cảm biến nhiệt bị đứt dây"),
    ("",  "Cột Cách khắc phục",        "Hướng dẫn kỹ thuật viên xử lý. VD: Thay cảm biến nhiệt NTC model XYZ"),
    ("",  "",  ""),
    ("",  "📌 VÍ DỤ LINH KIỆN",  ""),
    ("",  "Cột Mã linh kiện",          "LK-001, PN-234, ... (mã nội bộ của phòng bảo hành)"),
    ("",  "Cột Tên linh kiện",         "VD: Bơm thoát nước, Cảm biến NTC, Ron cao su cửa"),
    ("",  "Cột Mô tả / Dùng cho lỗi",  "VD: Dùng cho lỗi E3 – bơm đẩy nước thải ra ngoài"),
    ("",  "Cột Link ảnh linh kiện",    "Dán URL ảnh linh kiện (nếu có). Có thể để trống và gắn qua nút 🖼 trong chatbot."),
]

for r_idx, (_, col_b, col_c) in enumerate(rows_guide, start=3):
    row = r_idx
    ws_guide.row_dimensions[row].height = 22
    cb = ws_guide.cell(row=row, column=2, value=col_b)
    cc = ws_guide.cell(row=row, column=3, value=col_c)
    cc.alignment = align("left", "center", wrap=True)
    cc.font = font(size=10, color="334155")
    if col_b.startswith("📌"):
        cb.fill = fill("1E40AF")
        cb.font = font(bold=True, color=C_WHITE, size=10)
        cb.alignment = align("left", "center")
        ws_guide.merge_cells(f"B{row}:C{row}")
        ws_guide.row_dimensions[row].height = 26
    elif col_b:
        cb.fill = fill(C_LIGHT_BLUE)
        cb.font = font(bold=True, size=10, color="1E3A8A")
        cb.alignment = align("left", "center", wrap=True)
    else:
        ws_guide.row_dimensions[row].height = 10


# ─── SHEET 2: SẢN PHẨM ─────────────────────────────────
ws_sp = wb.create_sheet("🛒 SẢN PHẨM")
ws_sp.sheet_view.showGridLines = False
ws_sp.freeze_panes = "A3"

# Thiết lập độ rộng cột
col_widths_sp = {
    "A": 5,   # STT
    "B": 22,  # Loại sản phẩm
    "C": 20,  # Tên Model ⭐
    "D": 45,  # Đặc điểm nổi bật
    "E": 45,  # Hướng dẫn sử dụng
    "F": 35,  # Link ảnh sản phẩm
    "G": 35,  # Link ảnh hướng dẫn SD
}
for col_letter, width in col_widths_sp.items():
    ws_sp.column_dimensions[col_letter].width = width

# Header row 1 – tiêu đề sheet
ws_sp.merge_cells("A1:G1")
c = ws_sp["A1"]
c.value = "🛒 DANH SÁCH SẢN PHẨM – CHATBOT RAPIDO BẢO HÀNH"
c.fill = fill(C_HEADER_BLUE)
c.font = Font(bold=True, color=C_WHITE, size=13, name="Segoe UI")
c.alignment = align("center", "center")
ws_sp.row_dimensions[1].height = 36

# Header row 2 – tên cột
headers_sp = [
    ("A2", "STT",           C_HEADER_BLUE),
    ("B2", "Loại Sản Phẩm", C_HEADER_BLUE),
    ("C2", "⭐ Tên Model\n(BẮT BUỘC)", C_HEADER_BLUE),
    ("D2", "Đặc Điểm Nổi Bật",         C_HEADER_BLUE),
    ("E2", "Hướng Dẫn Sử Dụng\n(mỗi bước xuống dòng Alt+Enter)", C_HEADER_BLUE),
    ("F2", "🖼 Link Ảnh Sản Phẩm\n(URL http://...)",              C_HEADER_BLUE),
    ("G2", "🖼 Link Ảnh Hướng Dẫn\n(URL http://...)",             C_HEADER_BLUE),
]
ws_sp.row_dimensions[2].height = 44
for addr, text, bg in headers_sp:
    style_header_cell(ws_sp[addr], bg, text)

# Dữ liệu mẫu
sample_products = [
    (1, "Máy rửa chén", "RI-2000ES",
     "Tự động rửa chén, 8 chương trình, sấy khô bằng nhiệt, tiết kiệm điện A++",
     "1. Sắp xếp chén vào giỏ đúng chiều\n2. Thêm nước rửa + muối làm mềm nước\n3. Chọn chương trình phù hợp\n4. Nhấn START",
     "https://example.com/anh-may-rua-chen.jpg",
     "https://example.com/huong-dan-su-dung.jpg"),
    (2, "Máy sấy quần áo", "RD-500T",
     "Sấy bơm nhiệt Heat Pump, 7kg, 15 chương trình, tiết kiệm điện, giảm nhăn vải",
     "1. Kiểm tra túi lọc xơ vải trước mỗi lần dùng\n2. Cho quần áo vào, đóng cửa\n3. Chọn chương trình theo loại vải\n4. Nhấn START",
     "https://example.com/anh-may-say.jpg",
     ""),
    (3, "Tủ lạnh", "RF-300BX",
     "Tủ lạnh Inverter 300L, ngăn đông trên, làm lạnh nhanh, khử mùi tự động",
     "1. Đặt tủ cách tường tối thiểu 10cm\n2. Chờ 2 tiếng trước khi bỏ thực phẩm vào\n3. Không nhồi quá đầy ngăn tủ",
     "",
     ""),
]

for i, row_data in enumerate(sample_products, start=3):
    bg = C_ROW_ALT if i % 2 == 1 else C_WHITE
    ws_sp.row_dimensions[i].height = 50
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_sp.cell(row=i, column=col_idx)
        wrap = col_idx in (4, 5, 6, 7)
        bold = col_idx == 1
        style_data_cell(cell, value, bg, bold=bold, wrap=wrap,
                        align_h="center" if col_idx == 1 else "left")

# Thêm 30 dòng trống để người dùng điền
for i in range(len(sample_products) + 3, len(sample_products) + 3 + 30):
    bg = C_ROW_ALT if i % 2 == 1 else C_WHITE
    ws_sp.row_dimensions[i].height = 40
    # STT tự động gợi ý
    stt_cell = ws_sp.cell(row=i, column=1)
    style_data_cell(stt_cell, i - 2, bg, align_h="center", color="94A3B8")
    for col_idx in range(2, 8):
        cell = ws_sp.cell(row=i, column=col_idx)
        wrap = col_idx in (4, 5, 6, 7)
        if col_idx == 3:  # Cột Tên Model – highlight
            cell.fill = fill("FFFBEB")
        else:
            cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = align("left", "center", wrap=wrap)
        cell.font = font(size=10, color="64748B")

# Ghi chú cột C
ws_sp["C2"].fill = fill("1D4ED8")
note_c = ws_sp.cell(row=2, column=3)
note_c.comment = None  # openpyxl comment đơn giản


# ─── SHEET 3: MÃ LỖI ────────────────────────────────────
ws_err = wb.create_sheet("⚠️ MÃ LỖI")
ws_err.sheet_view.showGridLines = False
ws_err.freeze_panes = "A3"

col_widths_err = {
    "A": 22,  # Tên Model
    "B": 12,  # Mã lỗi
    "C": 45,  # Nguyên nhân
    "D": 50,  # Cách khắc phục
}
for col_letter, width in col_widths_err.items():
    ws_err.column_dimensions[col_letter].width = width

ws_err.merge_cells("A1:D1")
c = ws_err["A1"]
c.value = "⚠️ DANH SÁCH MÃ LỖI – NHẬP ĐÚNG TÊN MODEL NHƯ SHEET SẢN PHẨM"
c.fill = fill(C_HEADER_RED)
c.font = Font(bold=True, color=C_WHITE, size=13, name="Segoe UI")
c.alignment = align("center", "center")
ws_err.row_dimensions[1].height = 36

headers_err = [
    ("A2", "⭐ Tên Model\n(phải khớp sheet Sản Phẩm)", C_HEADER_RED),
    ("B2", "Mã Lỗi\nVD: E1, F1, H01",                 C_HEADER_RED),
    ("C2", "Nguyên Nhân Lỗi",                           C_HEADER_RED),
    ("D2", "Cách Khắc Phục",                            C_HEADER_RED),
]
ws_err.row_dimensions[2].height = 44
for addr, text, bg in headers_err:
    style_header_cell(ws_err[addr], bg, text)

sample_errors = [
    ("RI-2000ES", "E1", "Cảm biến nhiệt độ nước bị lỗi",
     "Kiểm tra và thay thế cảm biến nhiệt NTC. Vệ sinh bộ lọc nước đầu vào."),
    ("RI-2000ES", "E3", "Bơm thoát nước bị hỏng, nước không thoát được",
     "Kiểm tra bộ lọc đáy máy, thông tắc ống thoát. Nếu bơm hỏng thì thay mới."),
    ("RI-2000ES", "E5", "Cửa máy chưa đóng kín hoặc chốt hỏng",
     "Kiểm tra ron cao su cửa, đảm bảo chốt cửa khóa đúng vị trí."),
    ("RD-500T",   "F1", "Túi lọc xơ vải bị tắc nghẽn",
     "Làm sạch túi lọc xơ vải ở mặt trước. Vệ sinh bộ lọc phụ dưới cùng."),
    ("RD-500T",   "F4", "Cảm biến độ ẩm bị bẩn hoặc hỏng",
     "Lau sạch 2 thanh cảm biến độ ẩm trong lồng bằng vải khô. Nếu vẫn lỗi thì thay cảm biến."),
    ("RF-300BX",  "H01", "Nhiệt độ ngăn mát không đạt, máy chạy liên tục",
     "Kiểm tra ron cửa, đảm bảo đóng kín. Kiểm tra quạt dàn lạnh có bị đóng tuyết không."),
]

for i, row_data in enumerate(sample_errors, start=3):
    bg = C_ROW_ALT2 if i % 2 == 1 else C_WHITE
    ws_err.row_dimensions[i].height = 40
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_err.cell(row=i, column=col_idx)
        wrap = col_idx in (3, 4)
        bold = col_idx == 1
        style_data_cell(cell, value, bg, bold=bold, wrap=wrap,
                        align_h="left")

for i in range(len(sample_errors) + 3, len(sample_errors) + 3 + 50):
    bg = C_ROW_ALT2 if i % 2 == 1 else C_WHITE
    ws_err.row_dimensions[i].height = 36
    for col_idx in range(1, 5):
        cell = ws_err.cell(row=i, column=col_idx)
        wrap = col_idx in (3, 4)
        if col_idx == 1:
            cell.fill = fill("FFF1F2")
        else:
            cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = align("left", "center", wrap=wrap)
        cell.font = font(size=10, color="64748B")


# ─── SHEET 4: LINH KIỆN ─────────────────────────────────
ws_part = wb.create_sheet("🔩 LINH KIỆN")
ws_part.sheet_view.showGridLines = False
ws_part.freeze_panes = "A3"

col_widths_part = {
    "A": 22,  # Tên Model
    "B": 14,  # Mã linh kiện
    "C": 32,  # Tên linh kiện
    "D": 45,  # Mô tả / Dùng cho lỗi
    "E": 35,  # Link ảnh linh kiện
}
for col_letter, width in col_widths_part.items():
    ws_part.column_dimensions[col_letter].width = width

ws_part.merge_cells("A1:E1")
c = ws_part["A1"]
c.value = "🔩 DANH SÁCH LINH KIỆN THAY THẾ – NHẬP ĐÚNG TÊN MODEL NHƯ SHEET SẢN PHẨM"
c.fill = fill(C_HEADER_GREEN)
c.font = Font(bold=True, color=C_WHITE, size=13, name="Segoe UI")
c.alignment = align("center", "center")
ws_part.row_dimensions[1].height = 36

headers_part = [
    ("A2", "⭐ Tên Model\n(phải khớp sheet Sản Phẩm)",   C_HEADER_GREEN),
    ("B2", "Mã Linh Kiện\nVD: LK-001",                   C_HEADER_GREEN),
    ("C2", "Tên Linh Kiện",                               C_HEADER_GREEN),
    ("D2", "Mô Tả / Dùng Cho Lỗi Nào",                    C_HEADER_GREEN),
    ("E2", "🖼 Link Ảnh Linh Kiện\n(URL http://...)",       C_HEADER_GREEN),
]
ws_part.row_dimensions[2].height = 44
for addr, text, bg in headers_part:
    style_header_cell(ws_part[addr], bg, text)

sample_parts = [
    ("RI-2000ES", "LK-001", "Cảm biến nhiệt NTC",     "Dùng cho lỗi E1 – cảm biến đo nhiệt độ nước rửa",    ""),
    ("RI-2000ES", "LK-002", "Bơm thoát nước",          "Dùng cho lỗi E3 – bơm đẩy nước thải ra ngoài",       ""),
    ("RI-2000ES", "LK-003", "Ron cao su cửa",           "Dùng cho lỗi E5 – gioăng cao su làm kín cửa máy",    ""),
    ("RD-500T",   "LK-101", "Túi lọc xơ vải",          "Lọc xơ trong quá trình sấy, vệ sinh sau mỗi lần dùng",""),
    ("RD-500T",   "LK-102", "Thanh cảm biến độ ẩm",    "Dùng cho lỗi F4 – cảm nhận độ ẩm để điều chỉnh sấy",  ""),
    ("RF-300BX",  "LK-201", "Ron cửa tủ lạnh",         "Dùng cho lỗi H01 – gioăng cao su làm kín cửa tủ lạnh",""),
]

for i, row_data in enumerate(sample_parts, start=3):
    bg = C_ROW_ALT3 if i % 2 == 1 else C_WHITE
    ws_part.row_dimensions[i].height = 40
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_part.cell(row=i, column=col_idx)
        wrap = col_idx in (4, 5)
        bold = col_idx == 1
        style_data_cell(cell, value, bg, bold=bold, wrap=wrap,
                        align_h="left")

for i in range(len(sample_parts) + 3, len(sample_parts) + 3 + 60):
    bg = C_ROW_ALT3 if i % 2 == 1 else C_WHITE
    ws_part.row_dimensions[i].height = 36
    for col_idx in range(1, 6):
        cell = ws_part.cell(row=i, column=col_idx)
        wrap = col_idx in (4, 5)
        if col_idx == 1:
            cell.fill = fill("ECFDF5")
        else:
            cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = align("left", "center", wrap=wrap)
        cell.font = font(size=10, color="64748B")

# ─── Lưu file ───────────────────────────────────────────
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "template_sanpham_rapido.xlsx")
wb.save(output_path)
print(f"✅ Đã tạo file: {output_path}")
